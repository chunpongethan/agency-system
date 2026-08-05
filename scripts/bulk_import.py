"""
Bulk import: onboard agents and clients from CSV or XLSX.

Maps columns -> models, validates hierarchy integrity (levels, upline one level
above, no cycles), and runs a dry-run by default. Pass --commit to write.

Usage:
    python scripts/bulk_import.py --agents agents.csv --clients clients.csv
    python scripts/bulk_import.py --agents data.xlsx --commit          # xlsx sheets
    python scripts/bulk_import.py --agents agents.csv --commit \
        --database-url postgresql+psycopg2://user:pass@host/db

CSV columns
    agents:  code,name,email,level,upline_code,role,password
    clients: ref,name,email,phone,risk_profile,agent_code
XLSX: sheets named "agents" and "clients" with the same headers.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.models.models import Base, Agent, Client, Role
from app.services import agent_service
from app.security import hash_password


# --------------------------------------------------------------------------- #
# Readers
# --------------------------------------------------------------------------- #
def _read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return [ {k.strip(): (v.strip() if v else v) for k, v in row.items()}
                 for row in csv.DictReader(fh) ]


def _read_xlsx(path: str, sheet: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        out.append({header[i]: (str(r[i]).strip() if r[i] is not None else None)
                    for i in range(len(header))})
    return out


def _load(path: str | None, sheet: str) -> list[dict]:
    if not path:
        return []
    if path.lower().endswith((".xlsx", ".xlsm")):
        return _read_xlsx(path, sheet)
    return _read_csv(path)


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
class ImportError_(Exception):
    pass


def import_agents(session, rows: list[dict]) -> tuple[list[Agent], list[str]]:
    """Validate + build Agent objects in hierarchy order. Returns (agents, errors)."""
    errors: list[str] = []
    # Resolve existing codes -> ids.
    code_to_id: dict[str, int] = {
        a.code: a.id for a in session.execute(select(Agent)).scalars()
    }
    # Sort by level so uplines are created before their downlines.
    try:
        rows = sorted(rows, key=lambda r: int(r["level"]))
    except (KeyError, ValueError):
        errors.append("every agent row needs an integer 'level'")
        return [], errors

    built: list[Agent] = []
    for i, row in enumerate(rows, start=2):  # row 1 is the header
        code = row.get("code")
        if not code:
            errors.append(f"row {i}: missing code")
            continue
        try:
            level = int(row["level"])
        except (KeyError, ValueError):
            errors.append(f"row {i} ({code}): invalid level")
            continue

        upline_code = row.get("upline_code") or None
        upline_id = code_to_id.get(upline_code) if upline_code else None
        if upline_code and upline_id is None:
            errors.append(f"row {i} ({code}): upline '{upline_code}' not found")
            continue

        try:
            agent_service.validate_agent(session, level, upline_id)
        except agent_service.ValidationError as exc:
            errors.append(f"row {i} ({code}): {exc}")
            continue

        role = (row.get("role") or "agent").lower()
        try:
            role_enum = Role(role)
        except ValueError:
            errors.append(f"row {i} ({code}): invalid role '{role}'")
            continue

        agent = Agent(
            code=code, name=row.get("name") or code,
            email=row.get("email") or f"{code}@example.com",
            level=level, upline_id=upline_id, role=role_enum,
            password_hash=hash_password(row["password"]) if row.get("password") else None,
        )
        session.add(agent)
        session.flush()  # assign id so downlines in this batch can reference it
        code_to_id[code] = agent.id
        built.append(agent)
    return built, errors


def import_clients(session, rows: list[dict]) -> tuple[list[Client], list[str]]:
    errors: list[str] = []
    code_to_id = {a.code: a.id for a in session.execute(select(Agent)).scalars()}
    built: list[Client] = []
    for i, row in enumerate(rows, start=2):
        ref = row.get("ref")
        agent_code = row.get("agent_code")
        if not ref:
            errors.append(f"row {i}: missing ref")
            continue
        agent_id = code_to_id.get(agent_code)
        if agent_id is None:
            errors.append(f"row {i} ({ref}): agent '{agent_code}' not found")
            continue
        client = Client(
            ref=ref, name=row.get("name") or ref,
            email=row.get("email") or None, phone=row.get("phone") or None,
            risk_profile=row.get("risk_profile") or None, agent_id=agent_id,
        )
        session.add(client)
        session.flush()
        built.append(client)
    return built, errors


def run(agents_path, clients_path, database_url, commit) -> int:
    engine = create_engine(database_url)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    agent_rows = _load(agents_path, "agents")
    client_rows = _load(clients_path or agents_path, "clients")

    agents, agent_errors = import_agents(session, agent_rows)
    clients, client_errors = import_clients(session, client_rows)
    errors = agent_errors + client_errors

    print(f"Parsed {len(agent_rows)} agent row(s), {len(client_rows)} client row(s).")
    print(f"Valid: {len(agents)} agent(s), {len(clients)} client(s).")
    if errors:
        print(f"\n{len(errors)} problem(s):")
        for e in errors:
            print(f"  - {e}")

    if errors:
        session.rollback()
        print("\nErrors present -> nothing committed (fix and retry).")
        session.close()
        return 1

    if commit:
        session.commit()
        print("\nCommitted.")
    else:
        session.rollback()
        print("\nDry run -> nothing committed. Re-run with --commit to write.")
    session.close()
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Bulk import agents and clients.")
    p.add_argument("--agents", required=True, help="CSV/XLSX with agents (and clients sheet)")
    p.add_argument("--clients", help="CSV with clients (omit if using an xlsx with a clients sheet)")
    p.add_argument("--database-url", default=os.getenv("DATABASE_URL", "sqlite:///./backend/agency.db"))
    p.add_argument("--commit", action="store_true", help="write to the DB (default: dry run)")
    args = p.parse_args()
    return run(args.agents, args.clients, args.database_url, args.commit)


if __name__ == "__main__":
    raise SystemExit(main())
