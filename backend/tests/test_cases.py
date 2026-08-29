"""
Case (sales pipeline) access model: agents self-serve the cases they're assigned
to; managers view *and edit* their downlines' cases via the visibility layer;
admins do everything.
"""
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi.testclient import TestClient

from app.models.models import Base, Agent, Role
from app.security import hash_password


@pytest.fixture
def client():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False},
                           poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False)
    s = Session()

    def mk(code, level, role, upline=None):
        a = Agent(code=code, name=code, email=f"{code}@x.com", level=level, role=role,
                  upline_id=(upline.id if upline else None), password_hash=hash_password("pw"))
        s.add(a); s.flush()
        return a

    adm = mk("ADM", 1, Role.ADMIN)
    top = mk("TOP", 1, Role.MANAGER)
    ax = mk("AX", 2, Role.AGENT, top)
    ay = mk("AY", 2, Role.AGENT, top)
    other = mk("OTH", 1, Role.AGENT)  # unrelated agent, outside TOP's subtree
    s.commit()
    ids = {"adm": adm.id, "top": top.id, "ax": ax.id, "ay": ay.id, "other": other.id}
    s.close()

    from app import main
    def override_get_db():
        db = Session()
        try:
            yield db
        finally:
            db.close()
    main.app.dependency_overrides[main.get_db] = override_get_db
    tc = TestClient(main.app)
    tc._ids = ids
    yield tc
    main.app.dependency_overrides.clear()


def auth(tc, code):
    r = tc.post("/auth/login", json={"username": code, "password": "pw"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def mk_case(tc, headers, lead_id, **kw):
    payload = {"prospect_name": "P", "lead_agent_id": lead_id}
    payload.update(kw)
    return tc.post("/cases", headers=headers, json=payload)


def test_agent_creates_and_sees_own_case(client):
    ids = client._ids
    ax = auth(client, "AX")
    r = mk_case(client, ax, ids["ax"])
    assert r.status_code == 200, r.text
    case = r.json()
    assert case["stage"] == "lead" and case["outcome"] == "open"
    rows = client.get("/cases", headers=ax).json()
    assert any(c["id"] == case["id"] for c in rows)


def test_agent_cannot_create_case_for_others(client):
    ids = client._ids
    ax = auth(client, "AX")
    # AX assigns only AY (AX not on the case) -> 403.
    assert mk_case(client, ax, ids["ay"]).status_code == 403


def test_peer_agent_cannot_see_or_edit(client):
    ids = client._ids
    ax, ay = auth(client, "AX"), auth(client, "AY")
    case = mk_case(client, ax, ids["ax"]).json()
    assert all(c["id"] != case["id"] for c in client.get("/cases", headers=ay).json())
    assert client.patch(f"/cases/{case['id']}", headers=ay, json={"stage": "m1"}).status_code == 403


def test_manager_sees_and_edits_downline_case(client):
    ids = client._ids
    ax, top = auth(client, "AX"), auth(client, "TOP")
    case = mk_case(client, ax, ids["ax"]).json()
    rows = client.get("/cases", headers=top).json()
    assert any(c["id"] == case["id"] for c in rows)          # manager VIEWS downline case
    r = client.patch(f"/cases/{case['id']}", headers=top, json={"stage": "m1"})
    assert r.status_code == 200 and r.json()["stage"] == "m1"  # ... and may edit it


def test_manager_cannot_edit_case_outside_subtree(client):
    ids = client._ids
    top, other = auth(client, "TOP"), auth(client, "OTH")
    # A case owned by an unrelated agent outside TOP's subtree stays off-limits.
    case = mk_case(client, other, ids["other"]).json()
    assert client.patch(f"/cases/{case['id']}", headers=top,
                        json={"stage": "m1"}).status_code == 403


def test_admin_sees_all_and_can_edit(client):
    ids = client._ids
    ax, adm = auth(client, "AX"), auth(client, "ADM")
    case = mk_case(client, ax, ids["ax"]).json()
    assert any(c["id"] == case["id"] for c in client.get("/cases", headers=adm).json())
    assert client.patch(f"/cases/{case['id']}", headers=adm,
                        json={"stage": "m2"}).status_code == 200


def test_stage_and_outcome_transitions(client):
    ids = client._ids
    ax = auth(client, "AX")
    case = mk_case(client, ax, ids["ax"]).json()
    r = client.patch(f"/cases/{case['id']}", headers=ax, json={"stage": "m2"})
    assert r.status_code == 200 and r.json()["stage"] == "m2"
    r = client.patch(f"/cases/{case['id']}", headers=ax, json={"outcome": "won"})
    assert r.status_code == 200
    assert r.json()["outcome"] == "won" and r.json()["closed_at"] is not None
    # An open-only board no longer shows a won case.
    open_rows = client.get("/cases?outcome=open", headers=ax).json()
    assert all(c["id"] != case["id"] for c in open_rows)


def test_unrelated_agent_isolated(client):
    ids = client._ids
    ax, other = auth(client, "AX"), auth(client, "OTH")
    case = mk_case(client, ax, ids["ax"]).json()
    assert all(c["id"] != case["id"] for c in client.get("/cases", headers=other).json())


# --- Batch import ------------------------------------------------------------
import io  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from app.services import lead_import  # noqa: E402

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _wb_bytes(rows, headers=None):
    """Build an .xlsx with the template's friendly headers (or a custom set)."""
    wb = Workbook(); ws = wb.active
    ws.append(headers or [label for _, label in lead_import.COLUMNS])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_import_template_download(client):
    adm = auth(client, "ADM")
    r = client.get("/cases/import-template", headers=adm)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active.title == "Leads"
    assert wb.active.max_row >= 3            # header + 2 example rows
    assert "說明 Instructions" in wb.sheetnames


def test_batch_import_valid_and_invalid(client):
    adm = auth(client, "ADM")
    rows = [
        ["Alice", "a@x.com", "111", "AX", "", "", "lead", "participating,medical", 500000, "call", "ref"],
        ["Bob", "", "222", "AY", "AX", "", "prospect", "分紅險", "", "", ""],   # zh case-type label
        ["BadAgent", "", "", "ZZZ", "", "", "lead", "", "", "", ""],           # unknown agent
        ["BadStage", "", "", "AX", "", "", "nope", "", "", "", ""],            # invalid stage
        ["", "", "", "AX", "", "", "lead", "", "", "", ""],                    # missing name
        ["BadType", "", "", "AX", "", "", "lead", "flying", "", "", ""],       # unknown case type
    ]
    r = client.post("/cases/import", headers=adm,
                    files={"file": ("leads.xlsx", _wb_bytes(rows), _XLSX)})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 2
    assert body["failed"] == 4
    assert body["total"] == 6
    assert len(body["errors"]) == 4

    listed = client.get("/cases", headers=adm).json()
    by_name = {c["prospect_name"]: c for c in listed}
    assert {"Alice", "Bob"} <= set(by_name)
    assert set(by_name["Alice"]["case_types"]) == {"participating", "medical"}
    assert by_name["Alice"]["expected_afyp"] == 500000
    assert by_name["Bob"]["case_types"] == ["participating"]   # 分紅險 -> participating
    # distinct refs were assigned
    assert by_name["Alice"]["ref"] != by_name["Bob"]["ref"]


def test_import_respects_visible_scope(client):
    top = auth(client, "TOP")
    rows = [
        ["Downline", "", "", "AX", "", "", "lead", "", "", "", ""],   # AX is in TOP's subtree
        ["Outside", "", "", "OTH", "", "", "lead", "", "", "", ""],   # OTH is not
    ]
    r = client.post("/cases/import", headers=top,
                    files={"file": ("l.xlsx", _wb_bytes(rows), _XLSX)})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 1 and body["failed"] == 1
    assert "OTH" in body["errors"][0]["error"]


def test_import_accepts_machine_key_headers(client):
    """Headers can be the raw field names, not just the friendly zh labels."""
    adm = auth(client, "ADM")
    headers = ["prospect_name", "lead_agent_code", "stage"]
    rows = [["Zoe", "AX", "m1"]]
    r = client.post("/cases/import", headers=adm,
                    files={"file": ("l.xlsx", _wb_bytes(rows, headers), _XLSX)})
    assert r.status_code == 200, r.text
    assert r.json()["created"] == 1
    zoe = next(c for c in client.get("/cases", headers=adm).json() if c["prospect_name"] == "Zoe")
    assert zoe["stage"] == "m1"


def test_import_rejects_missing_required_columns(client):
    adm = auth(client, "ADM")
    # No lead-agent column at all -> the whole file is rejected (422).
    data = _wb_bytes([["Alice"]], headers=["prospect_name"])
    r = client.post("/cases/import", headers=adm, files={"file": ("l.xlsx", data, _XLSX)})
    assert r.status_code == 422
