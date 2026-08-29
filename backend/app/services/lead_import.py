"""Batch lead (case) import from an Excel workbook.

Builds a downloadable template and parses an uploaded .xlsx into raw row dicts.
Business rules (agent resolution, scope, persistence) stay in the endpoint;
this module only knows the spreadsheet shape.
"""
from __future__ import annotations

import io
import re

# Valid pipeline stages and case-type keys (mirrors the enums / frontend labels).
STAGES = ("lead", "prospect", "m1", "m2", "m3")
CASE_TYPE_KEYS = ("property", "eam", "participating", "medical", "hk_identity", "education")

# Accept a Chinese label (or a loose alias) for each case-type key, so users can
# type 分紅險 instead of `participating`.
_CASE_TYPE_ALIASES = {
    "property": "property", "房產方案": "property", "房产方案": "property", "房產": "property",
    "eam": "eam",
    "participating": "participating", "分紅險": "participating", "分红险": "participating", "分紅": "participating",
    "medical": "medical", "醫療重疾": "medical", "医疗重疾": "medical", "醫療": "medical", "重疾": "medical",
    "hk_identity": "hk_identity", "香港身份": "hk_identity", "身份": "hk_identity",
    "education": "education", "教育升學": "education", "教育升学": "education", "教育": "education",
}

# Ordered template columns: (machine field, friendly zh header).
COLUMNS = [
    ("prospect_name", "客戶姓名 *"),
    ("email", "電郵"),
    ("phone", "電話"),
    ("lead_agent_code", "Lead 代理編號 *"),
    ("sdr_agent_code", "SDR 代理編號"),
    ("closer_agent_code", "Closing 代理編號"),
    ("stage", "階段"),
    ("case_types", "個案類別"),
    ("expected_afyp", "預計AFYP"),
    ("follow_up", "跟進事項"),
    ("notes", "備註"),
]

# Header text a column may carry -> machine field. Matched case-insensitively
# against the first row, so the template's friendly headers and the raw machine
# keys both resolve.
_HEADER_ALIASES = {
    "prospect_name": {"prospect_name", "客戶姓名", "客户姓名", "姓名", "name", "prospect", "潛在客戶"},
    "email": {"email", "電郵", "电邮", "郵箱", "電子郵件", "e-mail"},
    "phone": {"phone", "電話", "电话", "手機", "聯絡電話", "tel", "mobile"},
    "lead_agent_code": {"lead_agent_code", "lead", "lead 代理編號", "lead代理編號", "lead代理", "lead agent", "跟進代理"},
    "sdr_agent_code": {"sdr_agent_code", "sdr", "sdr 代理編號", "sdr代理編號", "sdr代理", "sdr agent"},
    "closer_agent_code": {"closer_agent_code", "closer", "closing", "closing 代理編號", "closing代理編號", "closer agent", "成交代理"},
    "stage": {"stage", "階段", "阶段", "狀態", "status"},
    "case_types": {"case_types", "個案類別", "个案类别", "類別", "類型", "types", "case type"},
    "expected_afyp": {"expected_afyp", "預計afyp", "预计afyp", "afyp", "預期afyp"},
    "follow_up": {"follow_up", "跟進事項", "跟进事项", "跟進", "next action"},
    "notes": {"notes", "備註", "备注", "remarks", "note"},
}


def _norm(s) -> str:
    # Drop the "*" required-marker the template puts on some headers so it still
    # matches its plain alias.
    return re.sub(r"\s+", " ", str(s if s is not None else "").replace("*", "").strip()).lower()


def _match_header(cell) -> str | None:
    n = _norm(cell)
    if not n:
        return None
    for field, aliases in _HEADER_ALIASES.items():
        if n == field or n in {_norm(a) for a in aliases}:
            return field
    return None


def normalize_case_types(raw) -> tuple[list[str], list[str]]:
    """Split a case-types cell into (recognised keys, unknown tokens)."""
    if raw is None:
        return [], []
    tokens = [t for t in re.split(r"[,;、，\s]+", str(raw).strip()) if t]
    keys: list[str] = []
    unknown: list[str] = []
    for tok in tokens:
        key = _CASE_TYPE_ALIASES.get(_norm(tok))
        if key:
            if key not in keys:
                keys.append(key)
        else:
            unknown.append(tok)
    return keys, unknown


def build_template() -> bytes:
    """A styled .xlsx with the expected columns, two example rows, and an
    instructions sheet listing the valid values."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.alignment = Alignment(vertical="center")
    examples = [
        ["陳大文", "tai.man@example.com", "91234567", "A001", "A002", "A004",
         "lead", "participating,medical", 500000, "下週致電約見", "朋友轉介"],
        ["李小明", "", "98765432", "A001", "", "",
         "prospect", "property", "", "已發送房產資料", ""],
    ]
    for row in examples:
        ws.append(row)
    for i, w in enumerate([14, 22, 14, 16, 16, 18, 10, 26, 12, 24, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ref = wb.create_sheet("說明 Instructions")
    ref.append(["欄位 Column", "說明 Notes"])
    ref["A1"].font = ref["B1"].font = Font(bold=True)
    for r in [
        ("客戶姓名 *", "必填。潛在客戶姓名。"),
        ("Lead 代理編號 *", "必填。負責跟進的代理編號（例：A001），須為系統內有效且您可見的代理。"),
        ("SDR / Closing 代理編號", "可留空。填代理編號。"),
        ("階段 stage", "可留空（預設 lead）。有效值：lead / prospect / m1 / m2 / m3。"),
        ("個案類別 case_types", "可留空或多選，以逗號分隔。有效值：property(房產方案)、eam(EAM)、"
                              "participating(分紅險)、medical(醫療重疾)、hk_identity(香港身份)、education(教育升學)。"),
        ("預計AFYP", "可留空。數字（例：500000）。"),
        ("電郵 / 電話 / 跟進事項 / 備註", "可留空。"),
    ]:
        ref.append(r)
    ref.column_dimensions["A"].width = 28
    ref.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_rows(data: bytes) -> list[dict]:
    """Parse the uploaded workbook into [{"row": <1-based row>, "values": {field: raw}}]
    for each non-empty data row. Raises ValueError if the file is unreadable or is
    missing the required columns."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - surfaced to the caller as a 422
        raise ValueError(f"not a readable .xlsx workbook ({e})")
    ws = wb.active
    if ws is None:
        raise ValueError("workbook has no active sheet")
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise ValueError("the sheet is empty")

    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header):
        field = _match_header(cell)
        if field and field not in col_field.values():
            col_field[idx] = field
    if not {"prospect_name", "lead_agent_code"} <= set(col_field.values()):
        raise ValueError("missing required columns 客戶姓名 / Lead 代理編號")

    out: list[dict] = []
    rownum = 1
    for cells in it:
        rownum += 1
        values: dict[str, object] = {}
        any_val = False
        for idx, field in col_field.items():
            v = cells[idx] if idx < len(cells) else None
            if v is not None and str(v).strip() != "":
                any_val = True
            values[field] = v
        if any_val:
            out.append({"row": rownum, "values": values})
    wb.close()
    return out
